from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from src.core.models import TableLayout, ConversionResult

def sanitize_for_excel(text: str) -> str:
    """
    Sanitizes text to prevent Excel Injection (CSV Injection).
    Prepends a single quote if the text starts with dangerous characters.
    """
    if not text:
        return text
    # Dangerous characters that can start a formula
    if text.startswith(('=', '+', '-', '@')):
        return "'" + text
    return text

def render_excel(layout: TableLayout) -> ConversionResult:
    """
    Converts a TableLayout into an Excel file (bytes) and a preview DataFrame.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Converted Table"
    
    # Common styles
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data for Preview (Sparse Matrix approach)
    # We use a dictionary ref for sparse population then convert to list of lists
    preview_data = {} # (row, col) -> text
    max_preview_row = 0
    max_preview_col = 0
    
    # Sort cells by row then col (Critical for correct merging order)
    sorted_cells = sorted(layout.cells, key=lambda x: (x.row_index, x.col_index))
    
    occupied = set()
    
    for cell in sorted_cells:
        # 0-based to 1-based logic
        r = cell.row_index + 1
        c = cell.col_index + 1
        
        # 1. Skip if already occupied (prevent MergedCell read-only error)
        if (r, c) in occupied:
            continue

        safe_text = sanitize_for_excel(cell.text)
        
        # Write Value
        curr_cell = ws.cell(row=r, column=c, value=safe_text)
        
        # Apply style
        curr_cell.border = thin_border
        curr_cell.alignment = center_align
        
        # Populate Preview Data (Limit to top 50 rows for performance)
        if cell.row_index < 50:
            preview_data[(cell.row_index, cell.col_index)] = safe_text
            max_preview_row = max(max_preview_row, cell.row_index)
            max_preview_col = max(max_preview_col, cell.col_index)
        
        # Merge if needed
        if cell.row_span > 1 or cell.col_span > 1:
            end_r = r + cell.row_span - 1
            end_c = c + cell.col_span - 1
            
            # Check for overlapping merges (Safe Merge)
            is_safe_merge = True
            for ir in range(r, end_r + 1):
                for ic in range(c, end_c + 1):
                    # Check if any cell in target range (except start) is already occupied
                    if (ir, ic) in occupied and (ir, ic) != (r, c):
                        is_safe_merge = False
                        break
                if not is_safe_merge: break
            
            if is_safe_merge:
                ws.merge_cells(start_row=r, start_column=c, end_row=end_r, end_column=end_c)
                # Mark as occupied
                for ir in range(r, end_r + 1):
                    for ic in range(c, end_c + 1):
                        occupied.add((ir, ic))
            else:
                # Merge Conflict: Fallback to single cell
                occupied.add((r, c))
        else:
            occupied.add((r, c))
    
    from openpyxl.utils import get_column_letter
    
    # Auto-adjust column widths (rough approximation)
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            # MergedCell safe guard
            if isinstance(cell, MergedCell):
                continue
                
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    # 1. Generate Excel Bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.getvalue()
    
    # 2. Generate Preview DataFrame
    # Construct dense matrix from sparse data
    # Dimensions: (max_row + 1) x (max_col + 1)
    # We cap max_row at 50, but max_col can be dynamic
    rows = []
    for r in range(min(max_preview_row + 1, 50)):
        row_data = []
        for c in range(max_preview_col + 1):
            row_data.append(preview_data.get((r, c), "")) # Empty string for missing cells
        rows.append(row_data)
        
    df_preview = pd.DataFrame(rows)
    # Optional: Set better column names if first row looks like header? 
    # For now, default integer columns are safer to represent raw structure.

    return ConversionResult(
        excel_bytes=excel_bytes,
        preview_df=df_preview,
        row_count=layout.max_rows
    )
